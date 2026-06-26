import io
import csv
from django.http import HttpResponse
from django.db.models import Count, Sum, Avg, Q
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from drf_spectacular.utils import extend_schema, OpenApiResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.people.models import Student, Teacher
from apps.enrollment.models import AdminEnrollment
from apps.finance.models import Invoice, Payment
from apps.evaluation.models import Grade, SemesterResult
from apps.lms.models import CourseSpace, StudentProgress
from apps.academic.models import AcademicYear
from apps.programs.models import Program


def _excel_header_style(ws, row, headers, fill_color='1E3A8A'):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    font = Font(color='FFFFFF', bold=True, size=10)
    border = Border(
        bottom=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
    )
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = max(len(header) + 4, 15)
    ws.row_dimensions[row].height = 22


@extend_schema(responses={200: OpenApiResponse(description='Export Excel des étudiants')})
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_students_excel(request):
    academic_year_id = request.query_params.get('academic_year')
    program_id = request.query_params.get('program')

    qs = Student.objects.filter(is_active=True).select_related('user', 'current_program')
    if academic_year_id:
        qs = qs.filter(current_year_id=academic_year_id)
    if program_id:
        qs = qs.filter(current_program_id=program_id)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Étudiants'

    headers = ['Matricule', 'Nom', 'Prénom', 'Email', 'Genre', 'Nationalité',
               'Programme', 'Niveau', 'Statut', 'Date inscription']
    _excel_header_style(ws, 1, headers)

    alt_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    for i, s in enumerate(qs, 2):
        row_data = [
            s.student_id, s.user.last_name, s.user.first_name, s.user.email,
            'M' if s.gender == 'M' else 'F',
            s.nationality, s.current_program.name if s.current_program else '—',
            f'L{s.current_level}', s.get_status_display(),
            s.created_at.strftime('%d/%m/%Y') if s.created_at else '—',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            if i % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = Alignment(vertical='center')

    # Auto-fit
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="etudiants.xlsx"'
    return response


@extend_schema(responses={200: OpenApiResponse(description='Export Excel des notes')})
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_grades_excel(request):
    session_id = request.query_params.get('session')
    ec_id = request.query_params.get('ec')

    qs = Grade.objects.filter(status='publiee').select_related('student__user', 'ec__ue')
    if session_id:
        qs = qs.filter(exam_session_id=session_id)
    if ec_id:
        qs = qs.filter(ec_id=ec_id)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Notes'

    headers = ['Matricule', 'Nom complet', 'EC', 'UE', 'Note CC', 'Note Examen', 'Note Finale', 'Absent', 'Statut']
    _excel_header_style(ws, 1, headers)

    green_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

    for i, g in enumerate(qs, 2):
        row_data = [
            g.student.student_id, g.student.user.get_full_name(),
            g.ec.code, g.ec.ue.code,
            float(g.cc_grade) if g.cc_grade else '—',
            float(g.exam_grade) if g.exam_grade else '—',
            float(g.final_grade) if g.final_grade else '—',
            'OUI' if g.is_absent else 'NON',
            g.get_status_display(),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            if col == 7 and isinstance(val, float):
                cell.fill = green_fill if val >= 10 else red_fill
                cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="notes.xlsx"'
    return response


@extend_schema(responses={200: OpenApiResponse(description='Export CSV des paiements')})
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_payments_csv(request):
    academic_year_id = request.query_params.get('academic_year')

    qs = Payment.objects.filter(status='valide').select_related('invoice__student__user', 'invoice__academic_year')
    if academic_year_id:
        qs = qs.filter(invoice__academic_year_id=academic_year_id)

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="paiements.csv"'
    response.write('\ufeff')  # BOM UTF-8

    writer = csv.writer(response, delimiter=';')
    writer.writerow(['N° Reçu', 'Étudiant', 'Matricule', 'Montant', 'Mode', 'Date', 'Référence'])

    for p in qs:
        writer.writerow([
            p.receipt_number,
            p.invoice.student.user.get_full_name(),
            p.invoice.student.student_id,
            float(p.amount),
            p.get_method_display(),
            p.paid_at.strftime('%d/%m/%Y %H:%M') if p.paid_at else '—',
            p.transaction_ref or '—',
        ])

    return response


@extend_schema(responses={200: OpenApiResponse(description='Rapport statistique global')})
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def global_report(request):
    academic_year_id = request.query_params.get('academic_year')

    year_filter = {}
    if academic_year_id:
        year_filter['academic_year_id'] = academic_year_id

    # Effectifs
    students_total = Student.objects.filter(is_active=True).count()
    students_by_gender = list(Student.objects.filter(is_active=True).values('gender').annotate(count=Count('id')))
    students_by_status = list(Student.objects.filter(is_active=True).values('status').annotate(count=Count('id')))
    students_by_program = list(
        Student.objects.filter(is_active=True, current_program__isnull=False)
        .values('current_program__name', 'current_program__code')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )

    # Inscriptions
    enrollments = AdminEnrollment.objects.filter(**year_filter)
    enrollments_stats = {
        'total': enrollments.count(),
        'validees': enrollments.filter(status='validee').count(),
        'en_attente': enrollments.filter(status='en_attente').count(),
    }

    # Finance
    invoices = Invoice.objects.filter(**year_filter)
    finance_stats = invoices.aggregate(
        total_invoiced=Sum('total_amount'),
        total_paid=Sum('paid_amount'),
        total_invoices=Count('id'),
        paid_invoices=Count('id', filter=Q(status='payee')),
    )
    finance_stats['collection_rate'] = round(
        (float(finance_stats['total_paid'] or 0) / float(finance_stats['total_invoiced'] or 1)) * 100, 1
    )

    # Résultats académiques
    results = SemesterResult.objects.filter(published=True)
    if academic_year_id:
        results = results.filter(semester__program__department__faculty__university__isnull=False)
    results_stats = results.aggregate(
        avg_grade=Avg('average'),
        total=Count('id'),
        admis=Count('id', filter=Q(decision='admis')),
        ajournes=Count('id', filter=Q(decision='ajourné')),
        redoublants=Count('id', filter=Q(decision='redoublant')),
    )
    if results_stats['total']:
        results_stats['success_rate'] = round((results_stats['admis'] / results_stats['total']) * 100, 1)
    else:
        results_stats['success_rate'] = 0

    # LMS
    lms_stats = {
        'total_spaces': CourseSpace.objects.filter(is_published=True).count(),
        'avg_completion': StudentProgress.objects.aggregate(avg=Avg('completion_rate'))['avg'] or 0,
    }

    return Response({
        'students': {
            'total': students_total,
            'by_gender': students_by_gender,
            'by_status': students_by_status,
            'by_program': students_by_program,
        },
        'enrollments': enrollments_stats,
        'finance': finance_stats,
        'results': results_stats,
        'lms': lms_stats,
    })
